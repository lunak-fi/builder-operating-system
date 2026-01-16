import openpyxl
import logging
from pathlib import Path
from typing import Tuple
import email
from email import policy
from email.parser import BytesParser
import chardet

logger = logging.getLogger(__name__)


class DocumentParserError(Exception):
    """Raised when document parsing fails"""
    pass


def parse_excel(file_path: str) -> Tuple[str, dict]:
    """
    Parse Excel file and extract basic metadata and content summary.

    Args:
        file_path: Path to the Excel file (.xlsx or .xls)

    Returns:
        Tuple of (text_summary, metadata_dict)

    Raises:
        DocumentParserError: If parsing fails
    """
    try:
        # Verify file exists
        if not Path(file_path).exists():
            raise DocumentParserError(f"File not found: {file_path}")

        # Verify it's an Excel file
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            raise DocumentParserError(f"File is not an Excel file: {file_path}")

        logger.info(f"Parsing Excel file: {file_path}")

        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        sheets_info = []
        text_parts = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Get sheet dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column

            sheets_info.append({
                "name": sheet_name,
                "rows": max_row,
                "cols": max_col
            })

            # Sample first few rows for text summary
            text_parts.append(f"--- Sheet: {sheet_name} ({max_row} rows x {max_col} cols) ---")

            # Extract first 5 rows as sample
            sample_rows = []
            for row_idx, row in enumerate(sheet.iter_rows(max_row=5, values_only=True), start=1):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                sample_rows.append(f"Row {row_idx}: {' | '.join(row_values)}")

            text_parts.append("\n".join(sample_rows))

        workbook.close()

        text_summary = "\n\n".join(text_parts)
        metadata = {
            "sheets": sheets_info,
            "total_sheets": len(sheets_info)
        }

        logger.info(f"Successfully parsed Excel with {len(sheets_info)} sheets")

        return text_summary, metadata

    except openpyxl.utils.exceptions.InvalidFileException as e:
        raise DocumentParserError(f"Invalid or corrupted Excel file: {str(e)}")
    except Exception as e:
        if isinstance(e, DocumentParserError):
            raise
        raise DocumentParserError(f"Unexpected error parsing Excel: {str(e)}")


def parse_text_file(file_path: str) -> Tuple[str, dict]:
    """
    Parse text file with automatic encoding detection.

    Args:
        file_path: Path to the text file (.txt, .md, etc.)

    Returns:
        Tuple of (file_content, metadata_dict)

    Raises:
        DocumentParserError: If parsing fails
    """
    try:
        # Verify file exists
        if not Path(file_path).exists():
            raise DocumentParserError(f"File not found: {file_path}")

        logger.info(f"Parsing text file: {file_path}")

        # Read file as bytes first to detect encoding
        with open(file_path, 'rb') as f:
            raw_data = f.read()

        # Detect encoding
        encoding_result = chardet.detect(raw_data)
        encoding = encoding_result['encoding'] or 'utf-8'

        # Decode with detected encoding
        try:
            content = raw_data.decode(encoding)
        except (UnicodeDecodeError, AttributeError):
            # Fallback to utf-8
            content = raw_data.decode('utf-8', errors='replace')
            encoding = 'utf-8 (with errors replaced)'

        # Count lines and characters
        lines = content.splitlines()
        line_count = len(lines)
        char_count = len(content)

        metadata = {
            "encoding": encoding,
            "lines": line_count,
            "characters": char_count
        }

        logger.info(f"Successfully parsed text file: {line_count} lines, {char_count} chars")

        return content, metadata

    except Exception as e:
        if isinstance(e, DocumentParserError):
            raise
        raise DocumentParserError(f"Unexpected error parsing text file: {str(e)}")


def parse_email(file_path: str) -> Tuple[str, dict]:
    """
    Parse email file (.eml) and extract headers and body.

    Args:
        file_path: Path to the .eml file

    Returns:
        Tuple of (formatted_email_text, metadata_dict)

    Raises:
        DocumentParserError: If parsing fails
    """
    try:
        # Verify file exists
        if not Path(file_path).exists():
            raise DocumentParserError(f"File not found: {file_path}")

        # Verify it's an email file
        if not file_path.lower().endswith('.eml'):
            raise DocumentParserError(f"File is not an .eml file: {file_path}")

        logger.info(f"Parsing email file: {file_path}")

        # Parse email
        with open(file_path, 'rb') as f:
            msg = BytesParser(policy=policy.default).parse(f)

        # Extract headers
        from_addr = msg.get('From', 'Unknown')
        to_addr = msg.get('To', 'Unknown')
        subject = msg.get('Subject', 'No Subject')
        date = msg.get('Date', 'Unknown')
        cc_addr = msg.get('Cc', '')

        # Extract body
        body = ""
        if msg.is_multipart():
            # Get text/plain parts
            for part in msg.walk():
                content_type = part.get_content_type()
                if content_type == 'text/plain':
                    try:
                        body += part.get_content()
                    except Exception:
                        pass
        else:
            # Single part message
            try:
                body = msg.get_content()
            except Exception:
                body = str(msg.get_payload(decode=True), errors='replace')

        # Format as text
        formatted_text = f"""--- Email ---
From: {from_addr}
To: {to_addr}
Subject: {subject}
Date: {date}
{f'Cc: {cc_addr}' if cc_addr else ''}

--- Body ---
{body}
"""

        metadata = {
            "from": from_addr,
            "to": to_addr,
            "subject": subject,
            "date": date,
            "cc": cc_addr,
            "has_attachments": any(part.get_content_disposition() == 'attachment' for part in msg.walk()) if msg.is_multipart() else False
        }

        logger.info(f"Successfully parsed email: {subject}")

        return formatted_text, metadata

    except Exception as e:
        if isinstance(e, DocumentParserError):
            raise
        raise DocumentParserError(f"Unexpected error parsing email: {str(e)}")


def parse_document(file_path: str, file_type: str) -> Tuple[str, dict]:
    """
    Dispatcher function to parse documents based on file type.

    Args:
        file_path: Path to the document file
        file_type: Type of document ('pdf', 'excel', 'text', 'email', 'other')

    Returns:
        Tuple of (extracted_text, metadata_dict)

    Raises:
        DocumentParserError: If parsing fails
    """
    # Import PDF extractor here to avoid circular imports
    from app.services.pdf_extractor import extract_text_from_pdf, PDFExtractionError

    try:
        file_extension = Path(file_path).suffix.lower()

        # PDF
        if file_type == 'offer_memo' or file_extension == '.pdf':
            try:
                text = extract_text_from_pdf(file_path)
                file_size = Path(file_path).stat().st_size
                metadata = {"file_type": "pdf", "file_size_bytes": file_size}
                return text, metadata
            except PDFExtractionError as e:
                raise DocumentParserError(f"PDF parsing failed: {str(e)}")

        # Excel
        elif file_type == 'financial_model' or file_extension in ['.xlsx', '.xls']:
            return parse_excel(file_path)

        # Text
        elif file_type == 'transcript' or file_extension in ['.txt', '.md']:
            return parse_text_file(file_path)

        # Email
        elif file_type == 'email' or file_extension == '.eml':
            return parse_email(file_path)

        # Other/Unknown
        else:
            # Try to parse as text file as fallback
            logger.warning(f"Unknown file type '{file_type}', attempting text parsing")
            return parse_text_file(file_path)

    except Exception as e:
        if isinstance(e, DocumentParserError):
            raise
        raise DocumentParserError(f"Document parsing failed: {str(e)}")
