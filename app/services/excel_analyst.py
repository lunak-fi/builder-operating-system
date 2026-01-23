"""
Excel Financial Analyst Service

Extracts structured financial metrics from Excel financial models.
Handles variations in sheet names, cell locations, and data formats.
"""

import openpyxl
import logging
import re
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional
from difflib import SequenceMatcher

logger = logging.getLogger(__name__)


class ExcelAnalystError(Exception):
    """Raised when Excel analysis fails"""
    pass


# Common sheet name patterns (with variations)
SHEET_PATTERNS = {
    "returns": ["returns", "investment returns", "inv returns", "return", "investor returns"],
    "sources_uses": ["sources & uses", "sources and uses", "s&u", "s & u", "sources uses", "sources/uses"],
    "cash_flow": ["cash flow", "cashflow", "proforma", "pro forma", "cash flows", "projections"],
    "overview": ["overview", "summary", "executive summary", "deal summary", "investment summary"]
}

# Metric search patterns - maps field names to search terms
METRIC_PATTERNS = {
    "levered_irr": [
        "levered irr", "leveraged irr", "lp irr", "net irr", "irr to equity",
        "projected lp irr", "projected irr net", "irr"
    ],
    "unlevered_irr": [
        "unlevered irr", "unleveraged irr", "gross irr", "project level irr",
        "projected irr gross"
    ],
    "equity_multiple": [
        "equity multiple", "equity mult", "moic", "multiple on invested capital",
        "projected lp em", "net em", "multiple"
    ],
    "equity_required": [
        "equity required", "equity investment", "required equity", "lp equity",
        "total equity", "sponsor equity", "gp equity"
    ],
    "total_project_cost": [
        "total project cost", "total cost", "project cost", "total development cost",
        "total uses", "total investment"
    ],
    "land_cost": [
        "land cost", "purchase price", "acquisition price", "acquisition cost",
        "site cost", "land acquisition"
    ],
    "hard_cost": [
        "hard cost", "hard costs", "construction cost", "construction costs",
        "development cost", "building cost", "renovation cost", "capex"
    ],
    "soft_cost": [
        "soft cost", "soft costs", "fees", "closing costs", "transaction costs"
    ],
    "loan_amount": [
        "loan amount", "debt", "loan", "debt amount", "senior debt", "financing"
    ],
    "dscr_at_stabilization": [
        "dscr", "debt service coverage ratio", "stabilized dscr", "debt coverage"
    ],
    "exit_cap_rate": [
        "exit cap rate", "exit cap", "terminal cap rate", "terminal cap",
        "reversion cap rate", "going out cap"
    ],
    "yield_on_cost": [
        "yield on cost", "yoc", "stabilized yoc", "stabilized yield"
    ],
    "hold_period_months": [
        "hold period", "investment period", "hold", "project duration"
    ],
    "interest_rate": [
        "interest rate", "loan rate", "debt rate", "rate"
    ],
    "ltv": [
        "ltv", "loan to value", "loan-to-value", "loan to value ratio"
    ]
}


def fuzzy_match_sheet_name(sheet_name: str, pattern_list: List[str], threshold: float = 0.7) -> bool:
    """
    Fuzzy match a sheet name against a list of patterns.

    Args:
        sheet_name: The actual sheet name from Excel
        pattern_list: List of expected patterns
        threshold: Similarity threshold (0.0 to 1.0)

    Returns:
        True if sheet name matches any pattern
    """
    sheet_lower = sheet_name.lower().strip()

    for pattern in pattern_list:
        # Exact match
        if sheet_lower == pattern:
            return True

        # Fuzzy match
        similarity = SequenceMatcher(None, sheet_lower, pattern).ratio()
        if similarity >= threshold:
            return True

        # Substring match
        if pattern in sheet_lower or sheet_lower in pattern:
            return True

    return False


def find_sheet_by_type(workbook: openpyxl.Workbook, sheet_type: str) -> Optional[str]:
    """
    Find a sheet matching a specific type (returns, sources_uses, etc.)

    Args:
        workbook: openpyxl Workbook object
        sheet_type: Type key from SHEET_PATTERNS

    Returns:
        Sheet name if found, None otherwise
    """
    if sheet_type not in SHEET_PATTERNS:
        return None

    patterns = SHEET_PATTERNS[sheet_type]

    for sheet_name in workbook.sheetnames:
        if fuzzy_match_sheet_name(sheet_name, patterns):
            logger.info(f"Found {sheet_type} sheet: '{sheet_name}'")
            return sheet_name

    return None


def parse_numeric_value(cell_value: Any) -> Optional[float]:
    """
    Parse a cell value as a number, handling various formats.

    Handles:
    - Plain numbers: 1685348
    - Currency: $1,685,348
    - Percentages: 19.6% (returns as decimal 0.196)
    - Multipliers: 1.73x
    - Negative numbers

    Args:
        cell_value: Raw cell value from openpyxl

    Returns:
        Parsed float value or None if not parseable
    """
    if cell_value is None:
        return None

    # Already a number
    if isinstance(cell_value, (int, float)):
        return float(cell_value)

    # Convert to string and clean
    value_str = str(cell_value).strip()

    if not value_str:
        return None

    try:
        # Check if it's a percentage
        is_percentage = '%' in value_str

        # Remove common formatting
        cleaned = value_str.replace('$', '').replace(',', '').replace(' ', '')
        cleaned = cleaned.replace('(', '-').replace(')', '')  # Handle negative (123) format
        cleaned = cleaned.replace('%', '').replace('x', '')  # Remove % and x symbols

        # Parse as float
        number = float(cleaned)

        # Convert percentage to decimal
        if is_percentage and number > 1:  # Only if > 1 to avoid double conversion
            number = number / 100.0

        return number

    except (ValueError, AttributeError):
        return None


def search_for_metric(sheet, metric_name: str, search_terms: List[str]) -> Optional[float]:
    """
    Search for a metric in a sheet by looking for label cells and extracting adjacent values.

    Strategy:
    1. Find cells containing search terms (case-insensitive)
    2. Check adjacent cells (right, below) for numeric values
    3. Return first valid value found
    4. Prioritize longer/more specific matches over shorter ones

    Args:
        sheet: openpyxl worksheet
        metric_name: Name of metric being searched (for logging)
        search_terms: List of possible labels for this metric (should be ordered by specificity)

    Returns:
        Extracted numeric value or None
    """
    # Sort search terms by length (descending) to match more specific terms first
    sorted_terms = sorted(search_terms, key=len, reverse=True)

    for row in sheet.iter_rows(max_row=min(sheet.max_row, 200), max_col=min(sheet.max_column, 20)):
        for cell_idx, cell in enumerate(row):
            if cell.value is None:
                continue

            cell_text = str(cell.value).lower().strip()

            # Check if this cell matches any search term (prioritize longer matches)
            for term in sorted_terms:
                if term.lower() in cell_text:
                    logger.debug(f"Found potential label for '{metric_name}' at {cell.coordinate}: '{cell.value}'")

                    # Check adjacent cells for value (check multiple positions)
                    candidates = []

                    # Check right (up to 5 columns to the right)
                    for offset in range(1, 6):
                        if cell_idx + offset < len(row):
                            right_cell = row[cell_idx + offset]
                            value = parse_numeric_value(right_cell.value)
                            if value is not None and value != 0:  # Skip zeros
                                candidates.append((f"right+{offset}", right_cell.coordinate, value))
                                break  # Found a value, stop searching right

                    # Below (next row, same column)
                    below_row_num = cell.row + 1
                    if below_row_num <= sheet.max_row:
                        below_cell = sheet.cell(row=below_row_num, column=cell.column)
                        value = parse_numeric_value(below_cell.value)
                        if value is not None and value != 0:
                            candidates.append(("below", below_cell.coordinate, value))

                    # Return first valid value found
                    if candidates:
                        direction, coord, value = candidates[0]
                        logger.info(f"Extracted {metric_name} = {value} from {coord} ({direction} of label)")
                        return value

    return None


def extract_from_sheet(sheet, metrics_to_extract: List[str]) -> Dict[str, float]:
    """
    Extract specified metrics from a single sheet.

    Args:
        sheet: openpyxl worksheet
        metrics_to_extract: List of metric field names to extract

    Returns:
        Dictionary of {field_name: value}
    """
    extracted = {}

    for metric_name in metrics_to_extract:
        if metric_name not in METRIC_PATTERNS:
            continue

        search_terms = METRIC_PATTERNS[metric_name]
        value = search_for_metric(sheet, metric_name, search_terms)

        if value is not None:
            extracted[metric_name] = value

    return extracted


def extract_hold_period_months(workbook: openpyxl.Workbook, sheet_names: Dict[str, str]) -> Optional[int]:
    """
    Extract hold period and convert to months if needed.

    Args:
        workbook: openpyxl Workbook
        sheet_names: Dictionary mapping sheet types to actual sheet names

    Returns:
        Hold period in months, or None if not found
    """
    # Search in Returns or Overview sheets
    for sheet_type in ["returns", "overview"]:
        if sheet_type not in sheet_names:
            continue

        sheet = workbook[sheet_names[sheet_type]]
        search_terms = METRIC_PATTERNS["hold_period_months"]

        # Search for hold period
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 100)):
            for cell_idx, cell in enumerate(row):
                if cell.value is None:
                    continue

                cell_text = str(cell.value).lower().strip()

                # Check if mentions hold period
                if any(term in cell_text for term in ["hold period", "investment period", "hold"]):
                    # Check adjacent cells
                    if cell_idx + 1 < len(row):
                        value_cell = row[cell_idx + 1]
                        value = parse_numeric_value(value_cell.value)

                        if value is not None:
                            # Determine if it's in years or months based on context
                            if "month" in cell_text:
                                return int(value)
                            elif "year" in cell_text:
                                return int(value * 12)
                            else:
                                # Assume years if ambiguous and < 20
                                if value < 20:
                                    logger.info(f"Assuming hold period {value} is in years, converting to months")
                                    return int(value * 12)
                                else:
                                    return int(value)

    return None


def analyze_financial_model(
    excel_path: str,
    focus_metrics: Optional[List[str]] = None
) -> Dict[str, Any]:
    """
    Main function to analyze an Excel financial model and extract metrics.

    Args:
        excel_path: Path to Excel file
        focus_metrics: Optional list of specific metrics to extract (default: all)

    Returns:
        Dictionary with same structure as PDF extraction:
        {
            "underwriting": {
                "levered_irr": 0.196,
                "equity_multiple": 1.73,
                ...
                "_confidence": {...},
                "_cell_references": {...}
            }
        }

    Raises:
        ExcelAnalystError: If analysis fails
    """
    try:
        # Verify file exists
        if not Path(excel_path).exists():
            raise ExcelAnalystError(f"File not found: {excel_path}")

        if not excel_path.lower().endswith(('.xlsx', '.xls')):
            raise ExcelAnalystError(f"File is not an Excel file: {excel_path}")

        logger.info(f"Analyzing Excel financial model: {excel_path}")

        # Load workbook (data_only=True to get calculated values, not formulas)
        workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

        # Step 1: Identify sheet structure
        logger.info(f"Found {len(workbook.sheetnames)} sheets: {workbook.sheetnames}")

        sheet_names = {}
        for sheet_type in SHEET_PATTERNS.keys():
            found_sheet = find_sheet_by_type(workbook, sheet_type)
            if found_sheet:
                sheet_names[sheet_type] = found_sheet

        if not sheet_names:
            logger.warning("No recognized sheet patterns found, will search all sheets")

        # Step 2: Determine which metrics to extract
        if focus_metrics is None:
            focus_metrics = list(METRIC_PATTERNS.keys())

        # Remove hold_period_months from general extraction (handled specially below)
        metrics_to_search = [m for m in focus_metrics if m != "hold_period_months"]

        # Step 3: Extract metrics from appropriate sheets
        underwriting_data = {}

        # Priority order: Returns -> Sources & Uses -> Cash Flow -> Overview -> All sheets
        search_order = ["returns", "sources_uses", "cash_flow", "overview"]

        for sheet_type in search_order:
            if sheet_type not in sheet_names:
                continue

            sheet_name = sheet_names[sheet_type]
            sheet = workbook[sheet_name]

            logger.info(f"Searching '{sheet_name}' sheet for metrics")
            extracted = extract_from_sheet(sheet, metrics_to_search)

            # Merge extracted data (don't override existing values)
            for key, value in extracted.items():
                if key not in underwriting_data:
                    underwriting_data[key] = value

        # If still missing metrics, search ALL sheets as fallback
        missing_metrics = [m for m in metrics_to_search if m not in underwriting_data]
        if missing_metrics:
            logger.info(f"Searching all sheets for remaining metrics: {missing_metrics}")
            for sheet_name in workbook.sheetnames:
                if sheet_name in sheet_names.values():
                    continue  # Already searched

                sheet = workbook[sheet_name]
                extracted = extract_from_sheet(sheet, missing_metrics)

                for key, value in extracted.items():
                    if key not in underwriting_data:
                        underwriting_data[key] = value

        # Step 4: Special handling for hold period (convert years to months)
        # This runs regardless of whether it's in focus_metrics
        hold_period = extract_hold_period_months(workbook, sheet_names)
        if hold_period:
            underwriting_data["hold_period_months"] = hold_period

        # Step 5: Post-process extracted values
        # Convert negative costs to positive (they're often shown as outflows)
        cost_fields = ["land_cost", "hard_cost", "soft_cost", "total_project_cost",
                      "equity_required", "loan_amount"]
        for field in cost_fields:
            if field in underwriting_data and underwriting_data[field] < 0:
                underwriting_data[field] = abs(underwriting_data[field])
                logger.info(f"Converted {field} from negative to positive: {underwriting_data[field]}")

        # Step 6: Validate and compute confidence scores
        confidence = {}
        for metric, value in underwriting_data.items():
            # Assign confidence based on whether value seems reasonable
            if metric == "levered_irr" and 0 < value < 1:
                confidence[metric] = 0.95
            elif metric == "equity_multiple" and 0.5 < value < 10:
                confidence[metric] = 0.95
            elif metric == "dscr_at_stabilization" and 0.5 < value < 5:
                confidence[metric] = 0.90
            else:
                confidence[metric] = 0.85  # Default confidence

        workbook.close()

        # Step 6: Build result in same format as PDF extraction
        result = {
            "underwriting": underwriting_data,
            "_extraction_metadata": {
                "method": "excel",
                "sheets_found": sheet_names,
                "confidence": confidence
            }
        }

        metrics_found = len(underwriting_data)
        logger.info(f"Excel analysis complete: extracted {metrics_found} metrics")

        return result

    except openpyxl.utils.exceptions.InvalidFileException as e:
        raise ExcelAnalystError(f"Invalid or corrupted Excel file: {str(e)}")
    except Exception as e:
        if isinstance(e, ExcelAnalystError):
            raise
        raise ExcelAnalystError(f"Unexpected error during Excel analysis: {str(e)}")
